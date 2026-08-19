"""safe_xlsx.py — Excel (.xlsx) reader + L0 sanitizer.

Extracts visible cell values from .xlsx files using openpyxl, filtering out:
  - Hidden sheets (sheet_state = 'hidden' or 'veryHidden')
  - Hidden rows (row dimension hidden = True)
  - Near-white font color on near-white fill (invisible text)
  - Formula injection — cells starting with =, +, -, @ are defused

Then runs L0 sanitization on all extracted cell values.

Out of scope (v1):
  - Comments and notes
  - Named ranges that contain injection payloads
  - Conditional formatting that hides text
  - Password-protected workbooks

Usage (from Claude via Bash):
    python3 /path/to/safe_xlsx.py '/path/to/file.xlsx'

Exits non-zero on error. Output is sheet-labeled tab-separated text.
"""
import sys
import os

_THIS_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, _THIS_DIR)

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print(
        "ERROR: openpyxl not installed.\n"
        "Install with: pip install openpyxl",
        file=sys.stderr,
    )
    sys.exit(1)

from sanitize import sanitize, NO_CAP

try:
    from safe_input import scan_for_injection, provenance_route, resolve_desk
except ImportError:
    scan_for_injection = None; provenance_route = None; resolve_desk = None

_MAX_FILE_BYTES = 10_000_000   # 10 MB cap
_WHITE_THRESHOLD = 229         # RGB > this on all channels = near-white (out of 255)
_FORMULA_CHARS = frozenset("=+-@")


def _parse_color_value(color_value: str) -> tuple:
    """Parse openpyxl color value (AARRGGBB hex string) to (r, g, b) tuple or None."""
    if not color_value or color_value in ("00000000", "FF000000", None):
        return None
    try:
        # AARRGGBB format — skip alpha
        hex_str = color_value.lstrip("#")
        if len(hex_str) == 8:
            r = int(hex_str[2:4], 16)
            g = int(hex_str[4:6], 16)
            b = int(hex_str[6:8], 16)
            return (r, g, b)
        if len(hex_str) == 6:
            r = int(hex_str[0:2], 16)
            g = int(hex_str[2:4], 16)
            b = int(hex_str[4:6], 16)
            return (r, g, b)
    except (ValueError, TypeError):
        pass
    return None


def _is_white_color(color_value: str) -> bool:
    """Return True if this color value is near-white."""
    rgb = _parse_color_value(color_value)
    if rgb is None:
        return False
    r, g, b = rgb
    return r > _WHITE_THRESHOLD and g > _WHITE_THRESHOLD and b > _WHITE_THRESHOLD


def _is_hidden_cell(cell) -> bool:
    """Return True if cell text is effectively invisible (white on white)."""
    try:
        font = cell.font
        fill = cell.fill

        font_color = None
        if font and font.color:
            font_color = font.color.value if hasattr(font.color, 'value') else str(font.color)

        fill_color = None
        if fill and fill.fgColor:
            fill_color = fill.fgColor.value if hasattr(fill.fgColor, 'value') else str(fill.fgColor)

        font_white = _is_white_color(font_color) if font_color else False
        fill_white = _is_white_color(fill_color) if fill_color else False

        # Hidden if font is white (regardless of fill) — attacker doesn't need matching fill
        if font_white:
            return True

        # Also hidden if both are near-white (belt-and-suspenders)
        return False

    except Exception:
        return False


def _defuse_formula(value: str) -> str:
    """Strip leading formula injection char from cell value."""
    if value and value[0] in _FORMULA_CHARS:
        return value[1:].strip()
    return value


def extract_and_sanitize(path: str, desk: str = "root") -> str:
    """Extract visible cell values from .xlsx, filter hidden, run L0.

    Returns clean text organized by sheet. Raises RuntimeError on errors.
    """
    try:
        file_size = os.path.getsize(path)
    except OSError as e:
        raise RuntimeError(f"Cannot access file {path}: {e}") from e

    if file_size > _MAX_FILE_BYTES:
        raise RuntimeError(
            f"File too large: {file_size:,} bytes (max {_MAX_FILE_BYTES:,})."
        )

    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    except Exception as e:
        msg = str(e).lower()
        if "not a zip" in msg or "badzip" in msg or "corrupt" in msg:
            raise RuntimeError(f"Corrupt or invalid .xlsx file: {e}") from e
        raise RuntimeError(f"Failed to open .xlsx: {e}") from e

    output_sections = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Skip hidden and very-hidden sheets
        if ws.sheet_state in ("hidden", "veryHidden"):
            continue

        sheet_rows = []

        # Track hidden rows by row number
        hidden_rows = set()
        if hasattr(ws, 'row_dimensions'):
            for row_num, rd in ws.row_dimensions.items():
                if rd.hidden:
                    hidden_rows.add(row_num)

        for row in ws.iter_rows():
            if not row:
                continue

            row_num = row[0].row
            if row_num in hidden_rows:
                continue

            row_cells = []
            for cell in row:
                value = cell.value
                if value is None:
                    row_cells.append("")
                    continue

                str_value = str(value)

                # Skip hidden (white font) cells
                if _is_hidden_cell(cell):
                    row_cells.append("")
                    continue

                # Defuse formula injection
                defused = _defuse_formula(str_value)

                # L0 sanitization
                clean = sanitize(defused, max_len=NO_CAP)
                row_cells.append(clean)

            # Only include rows with at least one non-empty cell
            if any(c.strip() for c in row_cells):
                sheet_rows.append("\t".join(row_cells))

        if sheet_rows:
            output_sections.append(f"[Sheet: {sanitize(sheet_name, max_len=100)}]\n" +
                                    "\n".join(sheet_rows))

    wb.close()
    result = "\n\n".join(output_sections)

    # Heuristic injection scan — flags to stderr
    if scan_for_injection is not None:
        findings = scan_for_injection(result)
        if provenance_route is not None:
            provenance_route(desk, "file", result, item=path)   # on-path gate (W5): provenance + coverage breadcrumb + Sentinel verdict; format preserved in item=path
        if findings:
            print(f"[safe_xlsx] FLAGGED — {len(findings)} injection pattern(s) detected:", file=sys.stderr)
            for match, label in findings:
                print(f"  [{label}] \"{match}\"", file=sys.stderr)

    return result


if __name__ == "__main__":
    desk, rest = resolve_desk() if resolve_desk else ("root", sys.argv[1:])
    if not rest:
        print("Usage: python3 safe_xlsx.py [--desk <id>] '/path/to/file.xlsx'", file=sys.stderr)
        sys.exit(1)

    xlsx_path = rest[0]
    try:
        result = extract_and_sanitize(xlsx_path, desk=desk)
        print(result)
    except RuntimeError as e:
        print(f"ERROR: {e}", file=sys.stderr)
        sys.exit(1)
